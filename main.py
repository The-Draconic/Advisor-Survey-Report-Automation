import csv
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.utils import get_column_letter
import os

# COLORS
GRAY_FILL = 'B8B8B8'
GRAY_TEXT = '808080'
GREEN_FILL = '9EFC90'
ORANGE_FILL = 'F58505'
RED_FILL = 'FC9596'
RED_TEXT = '82030F'
YELLOW_FILL = 'FFE08A'
YELLOW_TEXT = '727801'

# BORDER ORIENTATIONS
# - 'box'       all sides
# - 'floor'     bottom
# - 'wall'      right side
# - 'corner'    bottom & right side
# - 'ham'       top and bottom
# - 'dog'       left and right
# - 'u'         all but top
# - 'd'         all but left

# Change the file extension to .xlsx.
# Args:
# - file (str): The name of the file.
# Returns:
# - str: The updated file name with the .xlsx extension.
def change_file_extension(file_path: str):
    if '.c' in file_path: 
        file_name = file_path.split('.c')[0]
        return f"{file_name}.xlsx"
    elif '.xlsx' in file_path: 
        return file_path

# Convert other responses to numerical values.
# Args:
# - response (str): The response to convert.
# Returns:
# - int: The converted numerical value.
def convert_other_response(response: str):
    if response == "Never" or response == "Not Convenient" or response == "None; I do no advanced preparation of my schedule.":
        return 1
    elif response == "Once" or response == "Convenient" or response == "Some; I pick out a few courses.":
        return 2
    elif response == "Twice" or response == "Very Convenient" or response == "A lot; I plan out most of my schedule in advance.":
        return 3
    elif response == "Three or more times":
        return 4
    else:
        return None

# Convert responses to numerical values.
# Args:
# - response (str): The response to convert.
# Returns:
# - int: The converted numerical value.
def convert_response(response: str):
    if response == "Strongly Disagree":
        return 1
    elif response == "Disagree":
        return 2
    elif response == "Agree":
        return 3
    elif response == "Strongly Agree":
        return 4
    else:
        return None

# Convert CSV file to XLSX format.
# Args:
# - file_path (str): The path of the CSV file to convert.
# Returns:
# - Workbook: The converted workbook.
def csv_to_xlsx(file_path: str):
    wb = Workbook()
    ws = wb.active
    with open(file_path, 'r') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, val in enumerate(row):
                ws.cell(row=r+1, column=c+1, value=val)  # Note: Excel uses 1-indexed rows and columns
    file_name = change_file_extension(file_path)
    wb.save(file_name)
    return wb

# Create collective statistics.
# Args:
# - file_path (str): The path of the file to create collective statistics for.
# - latest_year (str): The value of the latest year to append to the file name.
# Returns:
# - None
def create_collective_stats(file_path:str, latest_year: str):
    # Load the existing workbook
    existing_workbook = load_workbook(file_path)
    sheet_names = existing_workbook.sheetnames
    collective_sheet = sheet_names[-1]
    active_sheet = existing_workbook[collective_sheet]

    collective_stats_filename = f"All Adviser 3 Year Stats_{latest_year}.xlsx"
    if os.path.exists(collective_stats_filename): return collective_stats_filename

    # Create a new workbook for collective stats
    wb_collective_stats = Workbook()
    ws_collective_stats = wb_collective_stats.active

    # Collect info from file_path into data
    data = {}
    for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row-2, values_only=True):
        advisor = row[2]
        responses = row[3:]

        adivsor_dict = {f'Q{i}': response for i, response in enumerate(responses, start=1)}

        if advisor in data: data[advisor].append(adivsor_dict)
        else: data[advisor] = [adivsor_dict]

    # Calculate averages of all questions for all advisors
    question_sum = {}
    question_count = {}

    for advisor, surveys in data.items():
        for survey in surveys:
            average = 0
            count = 0
            for i in range(1, 18):
                question_key = f'Q{i}'
                response = survey.get(question_key)
                if question_key not in question_sum and question_key not in question_count:
                    question_sum[question_key] = 0
                    question_count[question_key] = 0
                if response is not None and int(response) != 0:
                    question_sum[question_key] += int(response)
                    question_count[question_key] += 1       
                    if i > 3:
                        average += int(response)
                        count += 1
            survey['Q18'] = format_average(average / count) if count != 0 else 0

    averages = {}
    for question, total in question_sum.items():
        count = question_count[question]
        if count > 0:
            averages[question] = format_average(total / count)
        else:
            averages[question] = 0.00  # Handle division by zero

    # Calculate the average of questions 4-17 for all advisors
    sum = 0
    count = 0
    for i in range(4, 18):
        sum += averages[f"Q{i}"]
        if averages[f"Q{i}"] != 0: count += 1
    averages["Q18"] = format_average(sum / count)

    # Set up styles
    neutral_style = NamedStyle(name="Neutral_Style")
    neutral_style.font = Font(bold=False, color=YELLOW_TEXT)
    neutral_style.fill = PatternFill(start_color=YELLOW_FILL, end_color=YELLOW_FILL, fill_type="solid")

    bad_style = NamedStyle(name="Bad_Style")
    bad_style.font = Font(bold=False, color=RED_TEXT)
    bad_style.fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")

    none_style = NamedStyle(name='None_Style')
    none_style.font = Font(bold=False, color=GRAY_TEXT)
    none_style.fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type='solid')

    # Headers
    headers = ["" , "# of responses", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10", "Q11", "Q12", "Q13", "Q14", "Q15", "Q16", "Q17", "Averages (#4-17)"]
    ws_collective_stats.append(headers)

    # Labels
    ws_collective_stats['V1'] = "Yellow cells = below average"
    ws_collective_stats['V1'].style = neutral_style
    ws_collective_stats['V2'] = "Red Cells = bottom 25%"
    ws_collective_stats['V2'].style = bad_style
    ws_collective_stats['V3'] = "Gray Cells = no responses"
    ws_collective_stats['V3'].style = none_style

    # Formatting
    for col in range(1, 23):
        if col == 1: width = 21
        elif col == 2: width = 13.29
        elif col == 20: width = 16.43
        else: width = 7.71
        col_letter = get_column_letter(col)
        ws_collective_stats.column_dimensions[col_letter].width = width

    # Iterate through data and populate the collective stats sheet
    for row, tuple in enumerate(list(data.items()), start=4):
        ws_collective_stats.cell(row=row, column=1, value=tuple[0]) # Advisor Name
        ws_collective_stats.cell(row=row, column=2, value=len(tuple[1])).alignment = Alignment(horizontal='center') # Total # of responses

        # Calculate averages for each question
        for col in range(3, 21):
            question_sum = 0
            question_count = 0
            for survey in tuple[1]:
                if survey[f"Q{col - 2}"] is None or survey[f"Q{col - 2}"] == 0: continue
                question_sum += int(survey[f"Q{col - 2}"])
                question_count += 1

            question_average = format_average(question_sum / question_count) if question_count > 0 else 0

            if question_average == 0: ws_collective_stats.cell(row=row, column=col, value=question_average).style = none_style
            elif is_bottom_25_percent(data, f"Q{col-2}", question_average): ws_collective_stats.cell(row=row, column=col, value=question_average).style = bad_style
            elif question_average < averages[f"Q{col-2}"]: ws_collective_stats.cell(row=row, column=col, value=question_average).style = neutral_style
            else: ws_collective_stats.cell(row=row, column=col, value=question_average)
        ws_collective_stats.cell(row=row, column=20).alignment = Alignment(horizontal='center')

    # Populate Averages row
    set_cell_properties(cell=ws_collective_stats.cell(row=3, column=1), value='Averages', bold=True, border='box')    
    set_cell_properties(cell=ws_collective_stats.cell(row=3, column=2), value='', border='box')    
    for col in range(3, 21):
        align = 'left' if col != 20 else 'center'
        set_cell_properties(cell=ws_collective_stats.cell(row=3, column=col), value=averages[f"Q{col-2}"], bold=True, border='box', alignment=Alignment(horizontal=align))
    
    # Apply autofilter to enable sorting in Excel
    ws_collective_stats.auto_filter.ref = "A3:T3"
    ws_collective_stats.freeze_panes = 'B4'

    # Save the collective stats workbook
    wb_collective_stats.save(collective_stats_filename)
    print(f"Collective stats saved to '{collective_stats_filename}'.")

    return collective_stats_filename

# Takes data from the two files put into the function related to the advisor and stores it into a separate excel sheet of said advisor's data.
# Args:
# - combined_file (str): The path of the 3 Year Combined Averages file.
# - collective_file (str): The path of the All Adviser 3 Year Stats file.
# - advisor (str): The name of the advisor whose file is being made with their survey data.
# Returns:
# - None
def duplicate_and_enter_data(combined_file:str, collective_file: str, advisor: str):
    # Define the workbooks of the two input files
    combined_wb = load_workbook(combined_file)
    collective_wb = load_workbook(collective_file)

    # Define the output_file, and it's workbook and worksheet
    output_file = f"{advisor}.xlsx"
    output_wb = Workbook()
    output_ws = output_wb.active

    # Define dictionary for advisor data
    advisor_data = {
        'averages': {},
        'gross_averages': {},
        'total_responses': 0,
        'surveys': []
    }

    # Find advisor data in combined workbook
    combined_sheet = combined_wb[combined_wb.sheetnames[-1]]  # Get the last sheet
    for row_index, row in enumerate(combined_sheet.iter_rows(min_row=2, max_row=combined_sheet.max_row, values_only=True), start=2):
        fill_color = combined_sheet.cell(row=row_index, column=1).fill
        if row[2] == advisor:
            advisor_data['surveys'].append((fill_color, row[3:]))  # Append row color and values from columns D-R
    
    # Get the colors of the three years
    legend = {}
    for index, sheet_name in enumerate(combined_wb.sheetnames[:3], start=0):
        print(f"Sheet Name: {sheet_name}")
        sheet = combined_wb[sheet_name]
        fill_color = sheet.cell(row=3, column=3).fill
        # print(f"Legend Fill Color: {fill_colors[index]}")
        legend[sheet_name] = fill_color
    # print(f"Legend: {legend}")

    # Place the legend
    for row_index, tuple in enumerate(list(legend.items()), start=5):
        # print(f"Tuple[1]: {tuple[1]}")
        year_color = tuple[1]
        year_start_color = year_color.start_color
        year_end_color = year_color.end_color
        year_style = year_color.fill_type
        year_fill = PatternFill(start_color=year_start_color, end_color=year_end_color, fill_type=year_style)
        set_cell_properties(output_ws.cell(row=row_index, column=5), value=tuple[0], fill=year_fill, alignment=Alignment(horizontal='center'))

    # Find advisor data in collective workbook
    collective_sheet = collective_wb.active
    for row in collective_sheet.iter_rows(min_row=3, max_row=collective_sheet.max_row, values_only=True):
        if row[0] == 'Averages':
            advisor_data['gross_averages'] = {f'Q{i}': value for i, value in enumerate(row[2:20], start=1)}
        if row[0] == advisor: 
            advisor_data['total_responses'] = row[1]
            advisor_data['averages'] = {f'Q{i}': value for i, value in enumerate(row[2:20], start=1)}
    print(f"Advisor: {advisor}\nAdvisor Data: {advisor_data}\nAdvisor Averages: {advisor_data['averages']}")

    # Format the output worksheet
    table_fill = PatternFill(start_color= GRAY_FILL, end_color= GRAY_FILL, fill_type='solid')
    question_fill = PatternFill(start_color=ORANGE_FILL, end_color=ORANGE_FILL, fill_type='solid')
    average_fill = PatternFill(start_color= GREEN_FILL, end_color= GREEN_FILL, fill_type='solid')

    # Averages Table
    output_ws.merge_cells('A1:C1')
    title_cell = output_ws['A1']
    set_cell_properties(title_cell, value='Averages', font_size=16, bold=True, alignment=Alignment(horizontal='center', wrap_text=True))
    output_ws.cell(row=1, column=3).border = Border(right=Side(style='medium'))
    for col in range(1, 4):
        output_ws.column_dimensions[get_column_letter(col)].width = 9.29
    output_ws.row_dimensions[2].height = 62.25
    top_text = ['Question#', 'Your Average Student Response', 'Average for all WWU Advisers']
    bottom_text = ['Averages', '=AVERAGE(B3:B17)', '=AVERAGE(C3:C17)']
    border_pattern = ['ham', 'ham', 'd']
    for col in range(1, 4): 
        set_cell_properties(output_ws.cell(row=2, column=col), value=top_text[col-1], alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), border=border_pattern[col-1])
        set_cell_properties(output_ws.cell(row=20, column=col), value=bottom_text[col-1], fill=table_fill, alignment=Alignment(horizontal='center'), border=border_pattern[col-1])
    for row in range(3, 20):
        border = 'floor' if row == 4 else None
        set_cell_properties(output_ws.cell(row=row, column=1), value=row-2, fill=table_fill, alignment=Alignment(horizontal='center'), border=border)            
        set_cell_properties(output_ws.cell(row=row, column=2), value=advisor_data['averages'][f'Q{row-2}'], fill=table_fill, alignment=Alignment(horizontal='center'), border='dog' if row != 4 else 'u')            
        set_cell_properties(output_ws.cell(row=row, column=3), value=advisor_data['gross_averages'][f'Q{row-2}'], fill=table_fill, alignment=Alignment(horizontal='center'), border=border)

    # Survey Questions Table
    output_ws.merge_cells('A22:C22')
    for merged_cells in output_ws['A22:C22']:
        # print(f"Cell: {merged_cells}")
        merged_cells[0].value = 'Questions 1-3'
        for cell in merged_cells: set_cell_properties(cell, font_size=21, alignment=Alignment(horizontal='center'), border='box')

    output_ws.merge_cells('D22:Q22')
    for merged_cells in output_ws['D22:Q22']:
        merged_cells[0].value = 'Questions 4-17'
        for cell in merged_cells: set_cell_properties(cell, font_size=21, alignment=Alignment(horizontal='center'), border='box')

    for row_index, survey in enumerate(advisor_data['surveys'], start=24):
        row_fill_color = survey[0]
        for col_index, data in enumerate(survey[1], start=1):
            cell_fill = PatternFill(start_color=row_fill_color.start_color, end_color=row_fill_color.end_color, fill_type=row_fill_color.fill_type)
            set_cell_properties(output_ws.cell(row=row_index, column=col_index), value=data, fill=cell_fill)
        output_ws.cell(row=row_index, column=3).border = Border(right=Side(style='medium'))
        output_ws.cell(row=row_index, column=17).border = Border(right=Side(style='medium'))
                
    # Apply formatting to header row (assuming it's row 21)
    for col in range(1, 18):
        set_cell_properties(output_ws.cell(row=23, column=col), value=f'Q #{col}', fill=question_fill, border='box')
            
    # Apply formatting to average row (assuming it's after the last row of survey data)
    max_row = output_ws.max_row
    for col in range(1, 18):
        set_cell_properties(output_ws.cell(row=max_row + 1, column=col), value=advisor_data['averages'][f'Q{col}'], fill=average_fill, alignment=Alignment(horizontal='center'), border='box')

    # Total Averages and Responses Section
    output_ws.column_dimensions['E'].width = 12.86
    output_ws.column_dimensions['F'].width = 13.57
    output_ws.column_dimensions['G'].width = 13
    summary_tuples = [('Total Average for questions 1-17', output_ws.cell(row=20, column=2).value), ('Total Average for questions 4-17', advisor_data['averages']['Q18']), ('Total Responses', advisor_data['total_responses'])]
    for col in range(5, 8):
        for row in range(2, 4):
            set_cell_properties(output_ws.cell(row=row, column=col), value=summary_tuples[col-5][row-2], alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), border='box')

    # Save the output workbook
    output_wb.save(f'Individual Advisor Data/{output_file}')
    print(f"Duplicated Excel sheet with data saved to '{output_file}'.")    

# Edit the combined sheet with survey data.
# Args:
# - file_path (str): The path of the file to edit.
# - survey_data (dict): Survey data to be added to the sheet.
# - new_year (int): An integer representing the new year of data
# Returns:
# - None
def edit_combined_sheet(file_path: str, survey_data: dict, new_year: int):
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # Check to see if sheet is already in workbook
    new_year_sheet_name = new_year
    if new_year_sheet_name in sheet_names: return

    # Get oldest year sheet
    oldest_year_sheet = sheet_names[0]
    old_year_sheet = wb[oldest_year_sheet]
    
    # Extract headers from the old year sheet
    headers = ['Student Names', 'Student Emails', 'Advisor Names'] + survey_data['Questions']
    # print(f'Headers: {headers}')

    # Extract fill color and column widths from the old year sheet
    # old_year_fill = [cell.fill for cell in old_year_sheet[1]]
    old_year_column_widths = [old_year_sheet.column_dimensions[get_column_letter(i)].width for i in range(1, len(headers) + 1)]
    
    wb.remove(old_year_sheet)
    wb.remove(wb[sheet_names[-1]])

    # Create new sheet for the new year
    new_year_sheet = wb.create_sheet(title=new_year_sheet_name)

    # Copy column widths from old year sheet to new year sheet
    for i, width in enumerate(old_year_column_widths, start=1):
        new_year_sheet.column_dimensions[get_column_letter(i)].width = width

    # Write survey data to new sheet
    new_year_sheet.append(headers)
    for cell in new_year_sheet[1]:
        cell.fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid")    
    for advisor, surveys in survey_data.items():
        # print(f"Surveys: {surveys}")
        if advisor == 'Questions': continue
        for survey in surveys:
            row_data = [survey['name'], survey['email'], advisor]
            row_data.extend([survey[f'Q{i}'] for i in range(1, len(survey_data['Questions']) + 1)])
            new_year_sheet.append(row_data)
        # Get the fill color of the first cell in the row
        cell_fill = old_year_sheet.cell(row=2, column=2).fill
        start_color = cell_fill.start_color
        end_color = cell_fill.end_color
        fill_type = cell_fill.fill_type
        for row_index in range(2, new_year_sheet.max_row + 1):
            for col in range(1, len(headers) + 1):
                # Apply the fill color to the current cell in the combined sheet
                new_year_sheet.cell(row=row_index, column=col).fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)

    # Update the sheet names list
    sheet_names = wb.sheetnames
    print(sheet_names)

    # Create new combined sheet
    oldest_year = sheet_names[0]
    newest_year = new_year_sheet_name
    combined_sheet = wb.create_sheet(title=f"{oldest_year}-{newest_year}")
    combined_sheet.append(headers)
    for cell in combined_sheet[1]: cell.fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type='solid')

    for i, width in enumerate(old_year_column_widths, start=1):
        combined_sheet.column_dimensions[get_column_letter(i)].width = width

    # Copy data from each sheet to the combined sheet with fill colors
    curr_row_index = 2
    for sheet_name in sheet_names:
        current_sheet = wb[sheet_name]
        print(f"Current Sheet: {current_sheet}")
        print(f"Current Sheet Max Row: {current_sheet.max_row}")
        max_row = current_sheet.max_row-166 if sheet_name == "2023" else current_sheet.max_row
        max_col = 20 if int(sheet_name) > 2023 else 18
        for row_index, row in enumerate(current_sheet.iter_rows(min_row=2, max_row=max_row, values_only=True), start=2):
            row_values = [cell for cell in row[:max_col]]
            combined_sheet.append(row_values)
            for col in range(1, max_col+1):
                # Get the fill color of the first cell in the row
                cell_fill = current_sheet.cell(row=2, column=col).fill
                start_color = cell_fill.start_color
                end_color = cell_fill.end_color
                fill_type = cell_fill.fill_type
                # Apply the fill color to the current cell in the combined sheet
                combined_sheet.cell(row=curr_row_index, column=col).fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
            curr_row_index += 1

    # Calculate averages
    question_columns = [f'{get_column_letter(i)}' for i in range(4, max_col+1)]
    question_averages = ['', '', 'Averages'] + [f'=AVERAGE({column}2:{column}{combined_sheet.max_row})' for column in question_columns]

    # Append question averages to the combined sheet
    combined_sheet.append(question_averages)

    # Calculate total average
    total_average_formula = f'=AVERAGE({question_columns[0]}{combined_sheet.max_row}:{question_columns[-1]}{combined_sheet.max_row})'
    combined_sheet.append(['', '', 'Total Average', total_average_formula])
    
    # Access the active sheet
    sheet_names = wb.sheetnames
    print(sheet_names)
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        wb.active = sheet
        # Apply autofilter to enable sorting in Excel
        sheet.auto_filter.ref = "A1:T1"
        sheet.freeze_panes = "A2"

    # Save workbook
    wb.save("3 Year Combined Averages.xlsx")
    print("Sheet editing complete.")

# Format floats into having up to 2 decimal numbers
# Args:
# - average (float): The float being formatted
# Returns:
# - average (float): The same float that has been reassigned its formatted form
def format_average(average: float):
    average = round(average, 2)
    average = float("{:.2f}".format(average))
    return average

# Get survey data from a file.
# Args:
# - file_path (str): The path of the file to get survey data from.
# Returns:
# - dict: Survey data extracted from the file.
def get_data(file_path: str):
    if not os.path.exists(file_path):
        source_file = input("What is the name of the file you want data from? ")
        survey_data = read_survey_data_new(source_file)

        with open(file_path, 'w') as json_file:
            json.dump(survey_data, json_file, indent=4)

        print(f"Survey data dumped into '{file_path}'.")
    else:
        with open(file_path, 'r') as json_file:
            survey_data = json.load(json_file)
    return survey_data

# See if the inputted average is in the bottom 25% of the list of averages
# Args:
# - data (dictionary): A dictionary of data of the advisors and their respective surveys.
# - question (str): A string representing what question is being surveyed
# - average (float): The target average being tested.
# Returns:
# - A boolean value, True if average is in the bottom 25%, False if not.
def is_bottom_25_percent(data: dict, question: str, average: float):
    # print(f"Question: {question}")
    # If average is 0, exempt it from bottom 25%
    if average == 0: 
        return False

    averages = []
    for advisor, surveys in data.items():
        sum = 0
        length = 0
        for survey in surveys:
            if survey[question] is None or survey[question] == 0: continue
            else: 
                sum += int(survey[question])
                length += 1
        question_average = sum / length if length != 0 else 0
        if question_average == 0: continue
        # if sum == 0 or length == 0: print(f"Advisor: {advisor}\tQuestion: {question}\tSum or Length is 0")
        averages.append(question_average)
    
    # Sort the list of averages in ascending order
    sorted_averages = sorted(averages)
    
    # Calculate the minimum and maximum values
    min_value = min(sorted_averages)
    max_value = max(sorted_averages)
    
    # Calculate the boundary for the bottom 25%
    boundary = min_value + 0.25 * (max_value - min_value)
    # if question == 'Q9': print(f"Averages: {sorted_averages}\nMinimum Value: {min_value}\nMaximum Value: {max_value}\nBoundary: {boundary}\nAverage: {average}\nIs Bottom 25?: {average <= boundary}")
    
    # Check if the given average falls within the bottom 25%
    return average <= boundary

# Read survey data from a CSV file. Compatible with sheets from 2023 and before.
# Args:
# - file_path (str): The path of the CSV file to read.
# Returns:
# - dict: Survey data extracted from the CSV file.
def read_survey_data(file_path: str='2023 Advisor Survey_January 25, 2024_16.19.csv'):
    workbook = csv_to_xlsx(file_path)

    # Load the Excel workbook
    sheet = workbook.active

    # Dictionary to store survey data
    survey_data = {}

    # Iterate through rows starting from row 4
    for row_number in range(4, sheet.max_row + 1):
        advisor_name =  sheet.cell(row=row_number, column=22).value
        first, last = advisor_name.split()
        advisor_name = f"{last}, {first}"
        if advisor_name:
            # Create a survey report dictionary
            survey_report = {
                'name': f"{sheet.cell(row=row_number, column=18).value} {sheet.cell(row=row_number, column=19).value}",
                'email': sheet.cell(row=row_number, column=20).value,
                'Q1': convert_other_response(sheet.cell(row=row_number, column=23).value),
                'Q2': convert_other_response(sheet.cell(row=row_number, column=24).value),
                'Q3': convert_other_response(sheet.cell(row=row_number, column=25).value),
                'Q4': convert_response(sheet.cell(row=row_number, column=26).value),
                'Q5': convert_response(sheet.cell(row=row_number, column=27).value),
                'Q6': convert_response(sheet.cell(row=row_number, column=28).value),
                'Q7': convert_response(sheet.cell(row=row_number, column=29).value),
                'Q8': convert_response(sheet.cell(row=row_number, column=30).value),
                'Q9': convert_response(sheet.cell(row=row_number, column=31).value),
                'Q10': convert_response(sheet.cell(row=row_number, column=32).value),
                'Q11': convert_response(sheet.cell(row=row_number, column=33).value),
                'Q12': convert_response(sheet.cell(row=row_number, column=34).value),
                'Q13': convert_response(sheet.cell(row=row_number, column=35).value),
                'Q14': convert_response(sheet.cell(row=row_number, column=36).value),
                'Q15': convert_response(sheet.cell(row=row_number, column=37).value),
            }

            # Append the survey report to the advisor's list in the survey_data dictionary
            if advisor_name in survey_data:
                survey_data[advisor_name].append(survey_report)
            else:
                survey_data[advisor_name] = [survey_report]

    return survey_data

# Read survey data from a CSV file. Compatible with the latest sheet (2024).
# Args:
# - file_path (str): The path of the CSV file to read.
# Returns:
# - dict: Survey data extracted from the CSV file.
def read_survey_data_new(file_path: str='2024 Advisor Survey_March 27, 2024_17.50.csv'):
    workbook = csv_to_xlsx(file_path)

    # Load the Excel workbook
    sheet = workbook.active

    # Dictionary to store survey data
    survey_data = {}

    # Get the Questions and add them to survey_data
    questions = []
    for col_number in range(18, 35):
        questions.append(sheet.cell(row=2, column=col_number).value)

    # Move the last two elements before the two elements before them
    questions[-4], questions[-3], questions[-2], questions[-1] = questions[-2], questions[-1], questions[-4], questions[-3]

    survey_data['Questions'] = questions
    print(f'Questions: {questions}')

    # Iterate through rows starting from row 4
    for row_number in range(4, sheet.max_row + 1):
        # If the data is part of a preview, skip it
        distribution = sheet.cell(row=row_number, column=16).value
        if distribution == 'preview': continue

        advisor_name =  sheet.cell(row=row_number, column=46).value
        first_name = sheet.cell(row=row_number, column=47).value
        last_name = sheet.cell(row=row_number, column=48).value
        name = f"{last_name}, {first_name}" if first_name not in [None, "", " "] or last_name not in [None, "", " "] else f"{advisor_name.split(' ')[1]}, {advisor_name.split(' ')[0]}"
        if name:
            # Create a survey report dictionary
            survey_report = {
                'name': sheet.cell(row=row_number, column=52).value,
                'email': sheet.cell(row=row_number, column=12).value,
                'Q1': convert_other_response(sheet.cell(row=row_number, column=18).value),
                'Q2': convert_other_response(sheet.cell(row=row_number, column=19).value),
                'Q3': convert_other_response(sheet.cell(row=row_number, column=20).value),
                'Q4': convert_response(sheet.cell(row=row_number, column=21).value),
                'Q5': convert_response(sheet.cell(row=row_number, column=22).value),
                'Q6': convert_response(sheet.cell(row=row_number, column=23).value),
                'Q7': convert_response(sheet.cell(row=row_number, column=24).value),
                'Q8': convert_response(sheet.cell(row=row_number, column=25).value),
                'Q9': convert_response(sheet.cell(row=row_number, column=26).value),
                'Q10': convert_response(sheet.cell(row=row_number, column=27).value),
                'Q11': convert_response(sheet.cell(row=row_number, column=28).value),
                'Q12': convert_response(sheet.cell(row=row_number, column=29).value),
                'Q13': convert_response(sheet.cell(row=row_number, column=30).value),
                'Q14': convert_response(sheet.cell(row=row_number, column=33).value),
                'Q15': convert_response(sheet.cell(row=row_number, column=34).value),
                'Q16': convert_response(sheet.cell(row=row_number, column=31).value),
                'Q17': convert_response(sheet.cell(row=row_number, column=32).value)
            }

            # Append the survey report to the advisor's list in the survey_data dictionary
            if name in survey_data:
                survey_data[name].append(survey_report)
            else:
                survey_data[name] = [survey_report]

    return survey_data

# Sets the properties of a cell
# Args:
# - cell, the cell being edited
# - value (str), the text being put into the cell
# - font_size (int), the size of the text
# - bold (bool), a boolean value for whether the text is bold or not
# - alignment (Alignment), the alignment that the cell is put in
# - border (str), a string representing the border formation it has
# Returns:
# - None
def set_cell_properties(cell, *, value: str = None, font_size: int= 11, fill: PatternFill= None, bold: bool= False, alignment: Alignment= None, border: str= None):
    side = Side(style='medium')
    if value: cell.value = value
    cell.font = Font(size=font_size, bold=bold)
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border == 'box': cell.border = Border(left=side, right=side, top=side, bottom=side)
    elif border == 'floor': cell.border = Border(bottom=side)
    elif border == 'wall': cell.border = Border(right=side)
    elif border == 'corner': cell.border = Border(right=side, bottom=side)
    elif border == 'ham': cell.border = Border(top=side, bottom=side)
    elif border == 'dog': cell.border = Border(left=side, right=side)
    elif border == 'u': cell.border = Border(left=side, right=side, bottom=side)
    elif border == 'd': cell.border = Border(right=side, top=side, bottom=side)

# Print the total number of surveys for each advisor.
# Args:
# - survey_data (dict): Survey data containing the advisor names and their respective surveys.
# Returns:
# - None
def print_survey_counts(survey_data):
    max = 0
    min = 0
    # Iterate through each advisor in the survey data
    for advisor, surveys in survey_data.items():
        # Keep track of maximum and minimum numbers of surveys amongst all advisors
        length = len(surveys)
        if max == 0 and min == 0:
            max = length
            min = length
        elif length > max: max = length
        elif length < min: min = length

        # Print the advisor's name and the total number of surveys
        print(f"Advisor: {advisor}, Total Surveys: {length}")
    print(f"Maximum Amount: {max}\tMinimum Amount: {min}")

if __name__ == "__main__":
    file = "3 Year Combined Averages.xlsx"
    new_year = input("What is the new year? (ex. 2024) ")
    survey_data = get_data(f'survey_data_{new_year}.json')
    # print_survey_counts(survey_data)
    edit_combined_sheet(file, survey_data, new_year)
    collective_stats = create_collective_stats(file, new_year)

    i = 0
    for advisor, surveys in survey_data.items():
        if advisor == "Questions": continue
        duplicate_and_enter_data(file, collective_stats, advisor)
        i += 1
        # if i == 1: break
