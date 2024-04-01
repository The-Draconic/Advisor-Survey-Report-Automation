import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

def format_excel_sheet(sheet):
    # Example formatting:
    sheet['A1'] = 'Sample Equation'
    sheet['B1'] = 'Formatted Cell'
    sheet['C1'] = 'Merged Cells'
    sheet['D1'] = 'Cell Borders'

    # Add an equation in cell A2
    sheet['A2'] = '=SUM(B2:B5)'

    # Format cell B2 with a yellow fill color
    sheet['B2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Merge cells C2 to D2
    sheet.merge_cells('C2:D2')
    sheet['C2'] = 'Merged Content'

    # Add borders to cells in column D
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

if True: # __name__ == "__test_main__":
    # Create a sample DataFrame
    data = {'Name': ['John', 'Alice', 'Bob', 'Charlie'],
            'Advisor': ['Advisor1', 'Advisor2', 'Advisor1', 'Advisor2']}
    df = pd.DataFrame(data)

    # Create an Excel writer and write the DataFrame to an Excel file
    with pd.ExcelWriter('formatted_output.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access the workbook and the active sheet
        workbook = writer.book
        sheet = workbook['Sheet1']

        # Call the function to apply formatting
        format_excel_sheet(sheet)

    print("Formatted Excel file generated successfully.")
